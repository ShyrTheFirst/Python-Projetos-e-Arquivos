from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import config


def username_finder():
    wb = load_workbook(filename = 'users.xlsx')
    ws = wb.active
    user = input("insert user ")
    find = False
    for index, cols in enumerate(ws.iter_cols(1)):
     for cell in cols:
      if cell.value == user:
       print("username found, insert password now")
       config.user_coordinate = (cell.row)
       find = True
    if find == False:
     print("Username not found")
     username_finder()
    else:
     password_finder()
    
def password_finder():
    wb = load_workbook(filename = 'users.xlsx')
    ws = wb.active
    pw = int(input("insert password "))
    find_pw = False
    for index, cols in enumerate(ws.iter_cols(2)):
     for cell in cols:
      if cell.value == pw:
       if cell.row == config.user_coordinate:
           config.pw_coordinate = (cell.row)
           if config.user_coordinate == config.pw_coordinate:
            print("Logged in")
           else:
            print("Password doesn't match")
            password_finder()
       else:
           pass
            
           find_pw = True
    if find_pw == False:
     print("Password doesn't match")
     password_finder()
    else:
     pass

username_finder()            
print("test worked!")


            
            

