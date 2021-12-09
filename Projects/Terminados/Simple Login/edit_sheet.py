from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import config as c

wb = load_workbook(filename = 'users.xlsx')
ws = wb.active
login_exist = False
def user_create():
    c.user = input("insert your user ")
    find = False
    for index, cols in enumerate(ws.iter_cols(1)):
     for cell in cols:
      if cell.value == c.user:
       print("username taken, try again")
       find = True
    if find == False:
     print("Username not found, You can use it! Create a password")
     password_create()
    else:
     user_create()
     
def password_create():
    print("Use only numbers for your password")
    c.password = int(input("Insert password "))
    user_and_password = (c.user,c.password)
    ws.append(user_and_password)
    




user_create()
print("Account created!")
wb.save('users.xlsx')
